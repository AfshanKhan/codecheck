"""Excel (.xlsx) reporter: a "Findings" sheet with AutoFilter and a frozen
header row, plus a "Summary" sheet breaking counts down by severity/check."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from codecheck.models import ReviewReport, Severity
from codecheck.reporters.glossary import format_ist, source_description, tier_description

_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _sanitize(value):
    """Neutralizes a leading formula-trigger character (=, +, -, @, tab, CR)
    against CSV/spreadsheet injection. The apostrophe is a visual marker;
    `_write_text_cell`'s Text number_format is what actually enforces it."""
    if isinstance(value, str) and value[:1] in _FORMULA_TRIGGERS:
        return "'" + value
    return value


def _write_text_cell(ws: Worksheet, row: int, col: int, value) -> None:
    cell = ws.cell(row=row, column=col, value=_sanitize(value))
    cell.number_format = "@"  # Text format -- never evaluated as a formula


_SEVERITY_FILL = {
    Severity.CRITICAL: PatternFill("solid", fgColor="B40000"),
    Severity.HIGH: PatternFill("solid", fgColor="C03B00"),
    Severity.MEDIUM: PatternFill("solid", fgColor="B46400"),
    Severity.LOW: PatternFill("solid", fgColor="1F5CA8"),
    Severity.INFO: PatternFill("solid", fgColor="606060"),
}
_HEADER_FILL = PatternFill("solid", fgColor="2F2F2F")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_WHITE_BOLD = Font(color="FFFFFF", bold=True)


def _autosize_columns(ws: Worksheet, widths: list[int]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def _write_summary_sheet(ws: Worksheet, report: ReviewReport) -> None:
    ws.title = "Summary"
    tiers_display = ", ".join(f"{t} ({tier_description(t)})" for t in report.tiers_run) or "-"
    # These can all be untrusted (from `codecheck render` on a hand-edited
    # report.json) -- text cells, not plain ws.append.
    text_rows = [
        ("Mode", report.mode),
        ("Base ref", report.base_ref or "-"),
        ("Head ref", report.head_ref or "-"),
        ("Tiers run", tiers_display),
    ]
    other_rows = [
        ("Generated at", format_ist(report.generated_at)),
        ("Files reviewed", len(report.files_reviewed)),
        ("Duration (s)", round(report.duration_seconds, 1)),
        ("Total findings", len(report.findings)),
        ("Compliance %", report.compliance_percentage()),
    ]
    ws.append(("Repo", None))
    _write_text_cell(ws, 1, 2, report.repo_path)
    for label, value in text_rows:
        row = ws.max_row + 1
        ws.append((label, None))
        _write_text_cell(ws, row, 2, value)
    for row in other_rows:
        ws.append(row)
    for cell in ws["A"][: len(text_rows) + len(other_rows) + 1]:
        cell.font = Font(bold=True)

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Severity", "Count"])
    for cell in ws[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for severity in (Severity.CRITICAL, Severity.HIGH, Severity.MEDIUM, Severity.LOW, Severity.INFO):
        count = sum(1 for f in report.findings if f.severity == severity)
        if count == 0:
            continue
        ws.append([severity.value.upper(), count])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.fill = _SEVERITY_FILL[severity]
        cell.font = _WHITE_BOLD

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Check", "Source", "Count", "Highest severity"])
    for cell in ws[header_row]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    by_check: dict[str, list] = {}
    for f in report.findings:
        by_check.setdefault(f.check_id, []).append(f)
    for check_id in sorted(by_check, key=lambda c: -len(by_check[c])):
        findings = by_check[check_id]
        worst = max(findings, key=lambda f: f.severity.rank).severity
        ws.append([None, None, len(findings), worst.value.upper()])
        row = ws.max_row
        _write_text_cell(ws, row, 1, check_id)
        _write_text_cell(ws, row, 2, findings[0].source)

    sources = sorted({f.source for f in report.findings})
    if sources:
        ws.append([])
        header_row = ws.max_row + 1
        ws.append(["Source", "What it is"])
        for cell in ws[header_row]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
        for s in sources:
            # source_description() echoes an unrecognized source back
            # unchanged -- both columns need the text-cell treatment.
            ws.append([None, None])
            row = ws.max_row
            _write_text_cell(ws, row, 1, s)
            _write_text_cell(ws, row, 2, source_description(s))

    if report.skipped:
        ws.append([])
        header_row = ws.max_row + 1
        ws.append(["Skipped"])
        ws.cell(row=header_row, column=1).font = Font(bold=True)
        for entry in report.skipped:
            ws.append([None])
            _write_text_cell(ws, ws.max_row, 1, entry)

    _autosize_columns(ws, [22, 60, 18, 18])


_FINDING_HEADERS = [
    "File", "Line", "Severity", "Check ID", "Source", "Tier",
    "Title", "Explanation", "Suggestion",
]


def _write_findings_sheet(ws: Worksheet, report: ReviewReport) -> None:
    ws.title = "Findings"
    ws.append(_FINDING_HEADERS)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    ordered = sorted(report.findings, key=lambda f: (-f.severity.rank, f.file, f.line_start))
    for f in ordered:
        ws.append([
            None,  # File -- written below via _write_text_cell
            f.line_start,
            f.severity.value.upper(),
            None,  # Check ID -- written below via _write_text_cell
            None,  # Source -- written below via _write_text_cell
            None,  # Tier -- written below via _write_text_cell
            None,  # Title -- written below via _write_text_cell
            None,  # Explanation -- written below via _write_text_cell
            None,  # Suggestion -- written below via _write_text_cell
        ])
        row = ws.max_row
        # All of these can be untrusted -- text-formatted so a leading
        # =/+/-/@ never gets evaluated as a formula.
        _write_text_cell(ws, row, 1, f.file)
        _write_text_cell(ws, row, 4, f.check_id)
        _write_text_cell(ws, row, 5, f.source)
        _write_text_cell(ws, row, 6, f.tier)
        _write_text_cell(ws, row, 7, f.title)
        _write_text_cell(ws, row, 8, f.explanation)
        _write_text_cell(ws, row, 9, f.suggestion or "")
        sev_cell = ws.cell(row=row, column=3)
        sev_cell.fill = _SEVERITY_FILL[f.severity]
        sev_cell.font = _WHITE_BOLD
        for col in (8, 9):  # Explanation / Suggestion -- long free text
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(_FINDING_HEADERS))}{ws.max_row}"
    _autosize_columns(ws, [45, 8, 10, 12, 10, 8, 40, 70, 50])


def render_xlsx(report: ReviewReport) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    _write_summary_sheet(wb.active, report)
    findings_ws = wb.create_sheet()
    _write_findings_sheet(findings_ws, report)
    return wb


def write_xlsx_report(report: ReviewReport, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    render_xlsx(report).save(str(output_path))
